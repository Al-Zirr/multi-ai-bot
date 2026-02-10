import io
import logging
import os
from pathlib import Path

from sqlalchemy import select

from bot.database import async_session
from bot.models.file import ProjectFile

logger = logging.getLogger(__name__)

# Supported file types and their MIME prefixes
SUPPORTED_TYPES = {
    ".pdf": "pdf",
    ".docx": "docx",
    ".doc": "docx",
    ".xlsx": "xlsx",
    ".xls": "xlsx",
    ".csv": "csv",
    ".txt": "txt",
    ".md": "txt",
    ".py": "txt",
    ".json": "txt",
    ".xml": "txt",
    ".html": "txt",
    ".log": "txt",
}


class FileService:
    def __init__(self, files_dir: str):
        self.files_dir = Path(files_dir)

    def _user_dir(self, user_id: int) -> Path:
        d = self.files_dir / str(user_id)
        d.mkdir(parents=True, exist_ok=True)
        return d

    async def save_file(
        self, user_id: int, filename: str, data: bytes
    ) -> ProjectFile | None:
        """Save file to disk and DB, extract text. Returns ProjectFile or None."""
        ext = Path(filename).suffix.lower()
        file_type = SUPPORTED_TYPES.get(ext)

        if not file_type:
            return None

        # Save to disk
        user_dir = self._user_dir(user_id)
        filepath = user_dir / filename
        # Avoid overwrite: add suffix
        counter = 1
        while filepath.exists():
            stem = Path(filename).stem
            filepath = user_dir / f"{stem}_{counter}{ext}"
            counter += 1

        filepath.write_bytes(data)

        # Extract text
        extracted = await self._extract_text(filepath, file_type)

        # Save to DB
        async with async_session() as session:
            pf = ProjectFile(
                user_id=user_id,
                filename=filepath.name,
                filepath=str(filepath),
                file_type=file_type,
                file_size=len(data),
                extracted_text=extracted,
            )
            session.add(pf)
            await session.commit()
            await session.refresh(pf)
            return pf

    async def get_user_files(self, user_id: int) -> list[ProjectFile]:
        async with async_session() as session:
            stmt = (
                select(ProjectFile)
                .where(ProjectFile.user_id == user_id)
                .order_by(ProjectFile.created_at.desc())
            )
            result = await session.execute(stmt)
            return list(result.scalars().all())

    async def get_file_by_id(self, file_id: int) -> ProjectFile | None:
        async with async_session() as session:
            return await session.get(ProjectFile, file_id)

    async def _extract_text(self, filepath: Path, file_type: str) -> str:
        """Extract text from file based on type."""
        try:
            if file_type == "pdf":
                return self._extract_pdf(filepath)
            elif file_type == "docx":
                return self._extract_docx(filepath)
            elif file_type == "xlsx":
                return self._extract_xlsx(filepath)
            elif file_type == "csv":
                return self._extract_csv(filepath)
            elif file_type == "txt":
                return filepath.read_text(encoding="utf-8", errors="replace")
        except Exception:
            logger.exception("Failed to extract text from %s", filepath)
            return ""
        return ""

    @staticmethod
    def _extract_pdf(filepath: Path) -> str:
        import fitz  # PyMuPDF

        text_parts = []
        with fitz.open(str(filepath)) as doc:
            for page in doc:
                text_parts.append(page.get_text())
        return "\n".join(text_parts)

    @staticmethod
    def _extract_docx(filepath: Path) -> str:
        from docx import Document

        doc = Document(str(filepath))
        return "\n".join(p.text for p in doc.paragraphs if p.text.strip())

    @staticmethod
    def _extract_xlsx(filepath: Path) -> str:
        from openpyxl import load_workbook

        wb = load_workbook(str(filepath), read_only=True, data_only=True)
        parts = []
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            parts.append(f"--- {sheet} ---")
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                parts.append("\t".join(cells))
        wb.close()
        return "\n".join(parts)

    @staticmethod
    def _extract_csv(filepath: Path) -> str:
        import pandas as pd

        df = pd.read_csv(str(filepath), nrows=10000)
        return df.to_string(index=False)

    @staticmethod
    def get_file_type_emoji(file_type: str) -> str:
        return ""
